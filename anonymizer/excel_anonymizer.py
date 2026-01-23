"""Excel en CSV anonimisatie module"""

import random
import string
import sys
import os
import csv
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from datetime import datetime, timedelta
import openpyxl
from openpyxl import load_workbook, Workbook

# Import security validators
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from utils.validators import escape_formula

try:
    from .reverse_anonymizer import AnonymizationMapping
    REVERSE_AVAILABLE = True
except ImportError:
    REVERSE_AVAILABLE = False


class ExcelColumnRule:
    """Excel kolom regel definitie"""

    def __init__(self, rule_dict: dict):
        self.id = rule_dict.get('id', '')
        self.column_name = rule_dict.get('columnName', '')
        self.column_type = rule_dict.get('columnType', 'text')  # text, date, number
        self.column_subtype = rule_dict.get('columnSubtype', None)  # supplier, artikelnaam, etc.
        self.anonymization_type = rule_dict.get('anonymizationType', 'replace')
        self.replace_with = rule_dict.get('replaceWith', '[ANONIEM]')
        self.date_offset_days = rule_dict.get('dateOffsetDays', 0)
        self.number_multiplier = rule_dict.get('numberMultiplier', 1.0)
        self.preserve_uniqueness = rule_dict.get('preserveUniqueness', False)
        self.reversible = rule_dict.get('reversible', False)  # Nieuwe parameter voor reversible kolommen

        # Nieuwe parameters voor prijsanonimisatie
        self.price_strategy = rule_dict.get('priceStrategy', 'fixed_multiplier')
        self.random_range_percent = rule_dict.get('randomRangePercent', 10)


class ExcelLogEntry:
    """Log entry voor Excel anonimisatie"""

    def __init__(self, rule_id: str, original: str, column: str, replaced: str, source_file: str = ""):
        self.rule_id = rule_id
        self.original = original
        self.column = column
        self.replaced = replaced
        self.source_file = source_file

    def to_dict(self) -> dict:
        return {
            'ruleId': self.rule_id,
            'originalTermDisplay': self.original,
            'appliedPattern': self.column,
            'replacedWith': self.replaced,
            'count': 1,
            'sourceFileName': self.source_file
        }


class ExcelAnonymizer:
    """Excel anonimisatie processor"""

    @staticmethod
    def generate_jabber(length: int = 8) -> str:
        """Genereer random string (jabber)"""
        chars = string.ascii_letters + string.digits
        return ''.join(random.choice(chars) for _ in range(length))

    @staticmethod
    def shuffle_column(values: List[any]) -> List[any]:
        """
        Shuffle column values preserving statistical distribution.

        Args:
            values: List of column values

        Returns:
            Shuffled list (same values, different order)

        Algorithm:
            1. Create copy of values list
            2. Use random.shuffle() for randomization
            3. Return shuffled copy

        Why shuffle preserves stats:
            - Mean: Sum unchanged, count unchanged → mean unchanged
            - Stddev: Variance based on value spread, not order → stddev unchanged
            - Min/max: Values unchanged → min/max unchanged
        """
        shuffled = values.copy()
        random.shuffle(shuffled)
        return shuffled

    @staticmethod
    def detect_column_type(values: List[any]) -> str:
        """
        Detect column type from sample values.

        Args:
            values: List of column values (sample or full)

        Returns:
            'number', 'date', or 'text'

        Algorithm:
            1. Filter out None/empty values
            2. Check first 20 non-empty values
            3. If >70% are numbers → 'number'
            4. Else if >70% are dates → 'date'
            5. Else → 'text'
        """
        # Filter out None and empty strings
        non_empty = [v for v in values[:20] if v not in (None, '', ' ')]
        if len(non_empty) == 0:
            return 'text'

        # Check numeric (int or float)
        numeric_count = sum(1 for v in non_empty if isinstance(v, (int, float)))
        if numeric_count / len(non_empty) > 0.7:
            return 'number'

        # Check date (datetime instances)
        date_count = sum(1 for v in non_empty if isinstance(v, datetime))
        if date_count / len(non_empty) > 0.7:
            return 'date'

        return 'text'

    @staticmethod
    def validate_rule_type_compatibility(rule: ExcelColumnRule, column_values: List[any]) -> Tuple[bool, str]:
        """
        Validate that rule transformation type is compatible with column data type.

        Args:
            rule: Excel column rule to validate
            column_values: List of column values

        Returns:
            Tuple of (is_valid, error_message)
            - is_valid: True if compatible, False otherwise
            - error_message: Empty string if valid, error description if invalid

        Type compatibility matrix:
            - number_multiply: requires 'number' column (not text/date)
            - date_offset: requires 'date' column (not text/number)
            - price_round: requires 'number' column (not text/date)
            - price_strategy: requires 'number' column (not text/date)
            - Other types: no restriction
        """
        detected_type = ExcelAnonymizer.detect_column_type(column_values)

        # Type compatibility matrix
        incompatible = {
            'number_multiply': ['text', 'date'],
            'date_offset': ['text', 'number'],
            'price_round': ['text', 'date'],
            'price_strategy': ['text', 'date']
        }

        # Check compatibility
        if rule.anonymization_type in incompatible:
            invalid_types = incompatible[rule.anonymization_type]
            if detected_type in invalid_types:
                return (False, f"Type mismatch: Cannot apply '{rule.anonymization_type}' to '{detected_type}' column '{rule.column_name}'")

        return (True, "")

    @staticmethod
    def save_workbook_as_csv(wb: Workbook, output_path: Path) -> None:
        """
        Save workbook as CSV with UTF-8 BOM for Excel compatibility.

        UTF-8 BOM (Byte Order Mark) signals Excel to use UTF-8 encoding,
        preventing corruption of special characters (é, ö, ü, etc.).

        Args:
            wb: Workbook to save
            output_path: Path to CSV file
        """
        # Get active sheet
        ws = wb.active

        # Extract all rows
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(row)

        # Write to CSV with UTF-8 BOM
        # utf-8-sig adds BOM automatically (EF BB BF hex)
        with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(rows)

    @staticmethod
    def process_excel_file(
        input_path: Path,
        output_path: Path,
        rules: List[ExcelColumnRule],
        preserve_headers: bool = True,
        mapping: Optional['AnonymizationMapping'] = None,
        reversible_mode: bool = False,
        sheet_names: Optional[List[str]] = None
    ) -> List[ExcelLogEntry]:
        """
        Verwerk een Excel bestand met sheet selectie

        Args:
            input_path: Pad naar input bestand
            output_path: Pad naar output bestand
            rules: List van Excel kolom regels
            preserve_headers: Behoud headers (eerste rij)
            mapping: Optional mapping voor reversible mode
            reversible_mode: Reversible mode actief
            sheet_names: List van sheet names om te processen (None = alle sheets)

        Returns:
            List van log entries

        Note:
            Excel formulas worden geconverteerd naar waarden tijdens processing.
            Originele formulas worden niet behouden.
        """
        # Laad workbook met alle mogelijke compatibiliteit opties
        wb = load_workbook(input_path, data_only=False, keep_vba=False)
        log_entries = []

        # Probeer metadata in te stellen om Excel warning te minimaliseren
        try:
            if hasattr(wb, 'properties'):
                wb.properties.creator = 'Microsoft Office User'
                wb.properties.lastModifiedBy = 'Microsoft Office User'
        except:
            pass  # Als dit faalt, geen probleem

        # Determine which sheets to process
        if sheet_names is None:
            # Process all sheets
            sheets_to_process = wb.sheetnames
        else:
            # Process only selected sheets (validate they exist)
            sheets_to_process = [name for name in sheet_names if name in wb.sheetnames]

        if len(sheets_to_process) == 0:
            raise ValueError("No valid sheets selected for processing")

        # Formula detection flag
        formula_warning_shown = False

        # Value mappings voor preserve uniqueness
        value_mappings: Dict[str, Dict[str, str]] = {}
        for rule in rules:
            if rule.preserve_uniqueness:
                value_mappings[rule.column_name] = {}

        # Price multiplier mappings voor reversible mode (veilige prijsanonimisatie)
        price_multipliers: Dict[str, float] = {}  # {original_price: multiplier}

        # Verwerk elke geselecteerde sheet
        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]

            # Check for formulas in first 10 rows (formula warning)
            if not formula_warning_shown:
                for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row)):
                    for cell in row:
                        if cell.data_type == 'f':  # Formula cell
                            print(f"[WARNING] Sheet '{sheet_name}' contains formulas. Formulas will be converted to values.")
                            formula_warning_shown = True
                            break
                    if formula_warning_shown:
                        break

            # Vind header rij (assumeer eerste rij)
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)

            # Maak column index mapping
            column_indices: Dict[str, int] = {}
            for idx, header in enumerate(headers, start=1):
                for rule in rules:
                    if rule.column_name == header:
                        column_indices[rule.column_name] = idx

            # Start rij (skip header als preserve_headers)
            start_row = 2 if preserve_headers else 1

            # Validate rule type compatibility before processing
            validated_rules = []
            for rule in rules:
                col_idx = column_indices.get(rule.column_name)
                if col_idx is None:
                    # Column not found in sheet - skip silently
                    continue

                # Extract column values for type detection
                column_values = []
                for row_idx in range(start_row, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    column_values.append(cell.value)

                # Validate type compatibility
                is_valid, error_msg = ExcelAnonymizer.validate_rule_type_compatibility(rule, column_values)
                if not is_valid:
                    # Log validation error and skip rule
                    print(f"[VALIDATION ERROR] {error_msg}")
                    continue

                # Rule is valid - add to validated list
                validated_rules.append(rule)

            # Use validated rules for processing (replaces original rules list for this sheet)
            rules = validated_rules

            # Pre-process shuffle rules (apply to entire column at once)
            shuffle_rules = [r for r in rules if r.anonymization_type == 'shuffle']
            for rule in shuffle_rules:
                col_idx = column_indices.get(rule.column_name)
                if col_idx is None:
                    continue

                # Extract all column values (skip header)
                column_values = []
                for row_idx in range(start_row, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    column_values.append(cell.value)

                # Shuffle values
                shuffled_values = ExcelAnonymizer.shuffle_column(column_values)

                # Apply shuffled values back to column
                for row_idx, new_value in enumerate(shuffled_values, start=start_row):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    original_value = cell.value

                    # Escape formula injection
                    if isinstance(new_value, str) and len(new_value) > 0 and new_value[0] in ('=', '+', '-', '@'):
                        print(f"FORMULA_ESCAPED cell={ws.title}!{cell.coordinate} original_value={new_value}")
                        new_value = escape_formula(new_value)

                    cell.value = new_value

                    # Log entry
                    if original_value is not None:
                        log_entries.append(ExcelLogEntry(
                            rule.id,
                            str(original_value),
                            rule.column_name,
                            str(new_value),
                            input_path.name
                        ))

            # Verwerk elke rij (skip shuffle rules - already processed above)
            non_shuffle_rules = [r for r in rules if r.anonymization_type != 'shuffle']
            for row_idx in range(start_row, ws.max_row + 1):
                for rule in non_shuffle_rules:
                    col_idx = column_indices.get(rule.column_name)
                    if col_idx is None:
                        continue

                    cell = ws.cell(row=row_idx, column=col_idx)
                    original_value = cell.value

                    if original_value is None:
                        continue

                    new_value = original_value
                    log_message = ""

                    # Text type
                    if rule.column_type == 'text':
                        if rule.anonymization_type == 'replace':
                            if rule.reversible:
                                # Reversible mode: maak unieke placeholders
                                if str(original_value) not in value_mappings.get(rule.column_name, {}):
                                    if rule.column_name not in value_mappings:
                                        value_mappings[rule.column_name] = {}
                                    unique_id = len(value_mappings[rule.column_name]) + 1
                                    # Maak placeholder gebaseerd op subtype of kolomnaam
                                    if rule.column_subtype:
                                        placeholder_prefix = rule.column_subtype.upper().replace(' ', '_')[:15]
                                    else:
                                        placeholder_prefix = rule.column_name.upper().replace(' ', '_')[:15]
                                    value_mappings[rule.column_name][str(original_value)] = f"[{placeholder_prefix}-{unique_id:03d}]"
                                new_value = value_mappings[rule.column_name][str(original_value)]

                                # Voeg toe aan mapping voor reversible mode
                                if mapping:
                                    mapping.add_mapping(str(original_value), str(new_value))
                            elif rule.preserve_uniqueness:
                                # Consistente unieke waarde (oude manier)
                                if str(original_value) not in value_mappings.get(rule.column_name, {}):
                                    if rule.column_name not in value_mappings:
                                        value_mappings[rule.column_name] = {}
                                    unique_id = len(value_mappings[rule.column_name]) + 1
                                    value_mappings[rule.column_name][str(original_value)] = f"{rule.replace_with}{unique_id}"
                                new_value = value_mappings[rule.column_name][str(original_value)]
                            else:
                                new_value = rule.replace_with

                        elif rule.anonymization_type == 'jabber':
                            if rule.reversible:
                                # Reversible mode: maak unieke placeholders
                                if str(original_value) not in value_mappings.get(rule.column_name, {}):
                                    if rule.column_name not in value_mappings:
                                        value_mappings[rule.column_name] = {}
                                    unique_id = len(value_mappings[rule.column_name]) + 1
                                    # Maak placeholder gebaseerd op subtype of kolomnaam
                                    if rule.column_subtype:
                                        placeholder_prefix = rule.column_subtype.upper().replace(' ', '_')[:15]
                                    else:
                                        placeholder_prefix = rule.column_name.upper().replace(' ', '_')[:15]
                                    value_mappings[rule.column_name][str(original_value)] = f"[{placeholder_prefix}-{unique_id:03d}]"
                                new_value = value_mappings[rule.column_name][str(original_value)]

                                # Voeg toe aan mapping voor reversible mode
                                if mapping:
                                    mapping.add_mapping(str(original_value), str(new_value))
                            elif rule.preserve_uniqueness:
                                if str(original_value) not in value_mappings.get(rule.column_name, {}):
                                    if rule.column_name not in value_mappings:
                                        value_mappings[rule.column_name] = {}
                                    value_mappings[rule.column_name][str(original_value)] = ExcelAnonymizer.generate_jabber(8)
                                new_value = value_mappings[rule.column_name][str(original_value)]
                            else:
                                new_value = ExcelAnonymizer.generate_jabber(8)

                    # Date type
                    elif rule.column_type == 'date':
                        if rule.anonymization_type == 'date_offset':
                            if isinstance(original_value, datetime):
                                new_value = original_value + timedelta(days=rule.date_offset_days)

                                # Voeg toe aan mapping voor reversible mode
                                if reversible_mode and mapping:
                                    mapping.add_mapping(str(original_value), str(new_value))

                    # Number type (prijzen)
                    elif rule.column_type == 'number':
                        if rule.anonymization_type == 'number_multiply':
                            if isinstance(original_value, (int, float)):
                                price_key = f"{rule.column_name}:{original_value}"

                                if rule.price_strategy == 'fixed_multiplier':
                                    # Strategie 1: Vaste multiplier (verhoudingen behouden voor BI)
                                    new_value = original_value * rule.number_multiplier
                                    if reversible_mode and mapping:
                                        mapping.add_mapping(str(original_value), str(new_value))

                                elif rule.price_strategy == 'random_per_price':
                                    # Strategie 2: Random per unieke prijs (geen verhoudingen, extra veilig)
                                    if price_key not in price_multipliers:
                                        # Genereer random multiplier tussen 0.15 en 0.65
                                        price_multipliers[price_key] = random.uniform(0.15, 0.65)
                                    new_value = original_value * price_multipliers[price_key]
                                    if reversible_mode and mapping:
                                        mapping.add_mapping(str(original_value), str(round(new_value, 2)))

                                elif rule.price_strategy == 'random_range':
                                    # Strategie 3: Random binnen range (verhoudingen ongeveer behouden)
                                    if price_key not in price_multipliers:
                                        # Random binnen ±percentage
                                        range_factor = rule.random_range_percent / 100.0
                                        min_mult = 1.0 - range_factor
                                        max_mult = 1.0 + range_factor
                                        price_multipliers[price_key] = random.uniform(min_mult, max_mult)
                                    new_value = original_value * price_multipliers[price_key]
                                    if reversible_mode and mapping:
                                        mapping.add_mapping(str(original_value), str(round(new_value, 2)))

                                # Round naar 2 decimalen voor prijzen
                                new_value = round(new_value, 2)

                    # Escape formula injection attacks before writing to cell
                    # Check if new_value starts with dangerous characters
                    if isinstance(new_value, str) and len(new_value) > 0 and new_value[0] in ('=', '+', '-', '@'):
                        # Log formula escaping for audit trail
                        print(f"FORMULA_ESCAPED cell={ws.title}!{cell.coordinate} original_value={new_value}")
                        new_value = escape_formula(new_value)

                    # Update cell
                    cell.value = new_value

                    # Log entry
                    log_entries.append(ExcelLogEntry(
                        rule.id,
                        str(original_value),
                        rule.column_name,
                        str(new_value),
                        input_path.name
                    ))

        # Final pass: Escape all remaining formulas in all cells (security hardening)
        # This catches any formulas that weren't processed by rules
        # Only escape string values - leave numbers and dates untouched
        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        original = cell.value
                        escaped = escape_formula(cell.value)
                        if original != escaped:
                            print(f"FORMULA_ESCAPED cell={ws.title}!{cell.coordinate} original_value={original}")
                            cell.value = escaped

        # Reset active cell naar A1 voor alle sheets (zodat Excel opent op A1)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Complete reset van sheet view om te forceren dat Excel opent op A1
            try:
                # Maak nieuwe sheet view aan
                from openpyxl.worksheet.views import SheetView, Selection
                sheet_view = SheetView()
                selection = Selection(activeCell='A1', sqref='A1')
                sheet_view.selection = [selection]
                sheet_view.topLeftCell = 'A1'
                ws.sheet_view = sheet_view

                # Verwijder freeze panes als die er zijn
                ws.freeze_panes = None
            except Exception as e:
                # Fallback naar oude methode
                ws.sheet_view.selection[0].activeCell = 'A1'
                ws.sheet_view.selection[0].sqref = 'A1'

        # Probeer calculation properties in te stellen (compatibiliteit met Excel)
        try:
            if hasattr(wb, 'calculation'):
                wb.calculation.calcMode = 'auto'
        except:
            pass

        # Save workbook - check if output should be CSV or XLSX
        if output_path.suffix.lower() == '.csv':
            # Save as CSV with UTF-8 BOM for Excel compatibility
            ExcelAnonymizer.save_workbook_as_csv(wb, output_path)
        else:
            # Save as Excel workbook with ISO dates voor betere compatibiliteit
            wb.save(output_path)

        return log_entries


def anonymize_excel(
    input_path: Path,
    output_path: Path,
    rules_data: List[dict],
    preserve_headers: bool = True
) -> Tuple[Path, List[dict]]:
    """
    Helper functie voor Excel anonimisatie

    Args:
        input_path: Input bestand pad
        output_path: Output bestand pad
        rules_data: List van regel dictionaries
        preserve_headers: Behoud headers

    Returns:
        Tuple van (output path, log entries als dictionaries)
    """
    rules = [ExcelColumnRule(r) for r in rules_data]
    logs = ExcelAnonymizer.process_excel_file(input_path, output_path, rules, preserve_headers)
    return output_path, [log.to_dict() for log in logs]
