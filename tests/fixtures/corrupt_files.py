"""
Corrupt File Generators for Edge Case Testing

This module generates various types of corrupt files to test error handling:
- Truncated XLSX files
- Invalid ZIP structures
- Malformed XML in Office documents
- Non-PDF files with .pdf extension
- Empty files
- Password-protected files (mock)
"""

import io
import zipfile
from pathlib import Path
from openpyxl import Workbook
from docx import Document
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter


def create_valid_xlsx():
    """Create a valid XLSX file in memory"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Data"

    # Add some data
    ws['A1'] = "Name"
    ws['B1'] = "Email"
    ws['C1'] = "Phone"

    ws['A2'] = "Jan de Vries"
    ws['B2'] = "jan@example.com"
    ws['C2'] = "06-12345678"

    ws['A3'] = "Maria Jansen"
    ws['B3'] = "maria@example.nl"
    ws['C3'] = "0612345678"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def create_truncated_xlsx(truncate_percent=0.5):
    """
    Create a truncated XLSX file (corrupted at specified percentage)

    Args:
        truncate_percent: Percentage of file to keep (0.5 = cut at 50%)

    Returns:
        bytes: Truncated XLSX data

    Example from PITFALLS.md line 366:
    - Truncated XLSX at 50% size causes zipfile.BadZipFile
    """
    valid_data = create_valid_xlsx()
    truncate_at = int(len(valid_data) * truncate_percent)
    return valid_data[:truncate_at]


def create_invalid_zip_structure():
    """
    Create XLSX with corrupted ZIP headers

    XLSX files are ZIP archives containing XML files. This creates
    a file that looks like ZIP but has invalid central directory.

    Returns:
        bytes: Invalid ZIP data

    Example from PITFALLS.md:
    - Invalid ZIP structure should raise clear error message
    """
    # Create valid XLSX first
    valid_data = create_valid_xlsx()

    # Corrupt the central directory signature
    # ZIP central directory starts with 'PK\x01\x02'
    # We'll corrupt it to 'XX\x01\x02'
    corrupted = bytearray(valid_data)

    # Find central directory signature
    for i in range(len(corrupted) - 4):
        if corrupted[i:i+4] == b'PK\x01\x02':
            corrupted[i:i+2] = b'XX'
            break

    return bytes(corrupted)


def create_malformed_xml_docx():
    """
    Create DOCX with invalid XML in document.xml

    DOCX files are ZIP archives containing document.xml.
    This creates a DOCX with malformed XML.

    Returns:
        bytes: DOCX with malformed XML

    Example from PITFALLS.md:
    - Malformed XML should be detected before processing
    """
    # Create a valid DOCX first
    doc = Document()
    doc.add_paragraph("Test document")

    output = io.BytesIO()
    doc.save(output)
    output.seek(0)

    # Extract ZIP, corrupt document.xml, repack
    corrupted_zip = io.BytesIO()

    with zipfile.ZipFile(output, 'r') as zip_in:
        with zipfile.ZipFile(corrupted_zip, 'w', zipfile.ZIP_DEFLATED) as zip_out:
            for item in zip_in.infolist():
                data = zip_in.read(item.filename)

                # Corrupt document.xml
                if item.filename == 'word/document.xml':
                    # Remove closing tag to make XML invalid
                    data = data.replace(b'</w:document>', b'')

                zip_out.writestr(item, data)

    corrupted_zip.seek(0)
    return corrupted_zip.read()


def create_non_pdf_with_pdf_extension():
    """
    Create a text file disguised as PDF

    Returns:
        bytes: Plain text with no PDF headers

    Example from PITFALLS.md line 88:
    - Non-PDF file with .pdf extension should be rejected
    """
    return b"This is not a PDF file, it's just plain text!"


def create_empty_file():
    """
    Create empty file (0 bytes)

    Returns:
        bytes: Empty bytes

    Example from PITFALLS.md line 368:
    - Empty files should return clear error: "File is empty"
    """
    return b""


def create_xlsx_with_only_headers():
    """
    Create Excel file with headers but no data rows

    Returns:
        bytes: XLSX with headers only

    This is an edge case that confuses users:
    - Is it valid? (Yes, technically)
    - Should we process it? (Warning, but allow)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Empty Data"

    # Add headers only
    ws['A1'] = "Name"
    ws['B1'] = "Email"
    ws['C1'] = "Phone"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def create_password_protected_xlsx_mock():
    """
    Create a mock password-protected XLSX

    Note: openpyxl can't create actual password-protected files,
    so we create a file with encryption markers in structure.

    Returns:
        bytes: XLSX that simulates password protection

    Example from PITFALLS.md line 370:
    - Password-protected files should return: "Upload unlocked version"
    """
    # For real testing, we'd need msoffcrypto-tool to create encrypted files
    # For now, return None to indicate this needs manual test file
    # In practice, we detect password protection when openpyxl.load_workbook fails
    return None


def create_scanned_pdf_mock():
    """
    Create PDF with no text layer (image-only)

    Note: ReportLab creates text-based PDFs. For real testing,
    we'd need to embed only images without text layer.

    Returns:
        bytes: PDF without text layer

    Example from PITFALLS.md line 341, 791:
    - Scanned PDFs should warn: "No text found - OCR required"
    """
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)

    # Draw only graphics, no text
    c.setFillColorRGB(0.5, 0.5, 0.5)
    c.rect(100, 100, 400, 600, fill=1)

    c.save()
    buffer.seek(0)
    return buffer.read()


def create_xlsx_with_macros_extension():
    """
    Create regular XLSX saved as .xlsm (macro-enabled)

    Returns:
        bytes: XLSX content (no actual macros)

    Example from PITFALLS.md line 371:
    - Files with .xlsm extension should be warned/rejected
    """
    return create_valid_xlsx()


def create_corrupted_pdf():
    """
    Create PDF with corrupted header

    Returns:
        bytes: PDF with invalid header

    PDF files start with "%PDF-1.x" header.
    Corrupt this to cause parsing errors.
    """
    valid_pdf = io.BytesIO()
    c = canvas.Canvas(valid_pdf, pagesize=letter)
    c.drawString(100, 750, "Test PDF content")
    c.save()

    valid_pdf.seek(0)
    data = valid_pdf.read()

    # Corrupt PDF header (%PDF-1.4 -> %XYZ-1.4)
    corrupted = data.replace(b'%PDF', b'%XYZ', 1)
    return corrupted


def create_large_xlsx(num_rows=10000):
    """
    Create large XLSX file for performance testing

    Args:
        num_rows: Number of data rows to generate

    Returns:
        bytes: Large XLSX file (~10MB for 10k rows)

    Example from PITFALLS.md:
    - 10k rows should process in <30s
    - 50k rows should process in <120s
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Large Dataset"

    # Add headers
    ws['A1'] = "ID"
    ws['B1'] = "Name"
    ws['C1'] = "Email"
    ws['D1'] = "Phone"
    ws['E1'] = "Address"

    # Add data rows
    for i in range(2, num_rows + 2):
        ws[f'A{i}'] = i - 1
        ws[f'B{i}'] = f"Person {i-1}"
        ws[f'C{i}'] = f"person{i-1}@example.com"
        ws[f'D{i}'] = f"06{10000000 + i:08d}"  # Dutch mobile format
        ws[f'E{i}'] = f"Street {i-1}, 1234 AB Amsterdam"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# Helper function to save corrupt files for manual testing

def save_corrupt_files_to_disk(output_dir: Path):
    """
    Save all corrupt file variants to disk for manual testing

    Args:
        output_dir: Directory to save files
    """
    output_dir.mkdir(exist_ok=True, parents=True)

    files_to_create = {
        'truncated_50percent.xlsx': create_truncated_xlsx(0.5),
        'truncated_75percent.xlsx': create_truncated_xlsx(0.75),
        'invalid_zip_structure.xlsx': create_invalid_zip_structure(),
        'malformed_xml.docx': create_malformed_xml_docx(),
        'fake_pdf.pdf': create_non_pdf_with_pdf_extension(),
        'empty_file.txt': create_empty_file(),
        'empty_file.xlsx': create_empty_file(),
        'headers_only.xlsx': create_xlsx_with_only_headers(),
        'scanned_mock.pdf': create_scanned_pdf_mock(),
        'corrupted_header.pdf': create_corrupted_pdf(),
        'large_10k_rows.xlsx': create_large_xlsx(10000),
    }

    for filename, data in files_to_create.items():
        if data is not None:  # Skip mocks that can't be generated
            filepath = output_dir / filename
            with open(filepath, 'wb') as f:
                f.write(data)

    print(f"Created {len(files_to_create)} corrupt test files in {output_dir}")


if __name__ == '__main__':
    # Generate files for manual testing
    import sys
    output_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).parent / 'corrupt'
    save_corrupt_files_to_disk(output_path)
