"""
Tests for CSV UTF-8 BOM encoding
"""

import pytest
import sys
import os
import tempfile
from pathlib import Path

# Add parent directory to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import Workbook
from anonymizer.excel_anonymizer import ExcelAnonymizer


class TestCSVEncoding:
    """Test CSV encoding with UTF-8 BOM"""

    def test_csv_has_utf8_bom(self):
        """Test that CSV export includes UTF-8 BOM"""
        # Create test workbook with special characters
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Description"])
        ws.append(["Café", "Müller"])
        ws.append(["José", "García"])
        ws.append(["naïve", "résumé"])

        # Save as CSV
        with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as f:
            output_path = Path(f.name)

        try:
            ExcelAnonymizer.save_workbook_as_csv(wb, output_path)

            # Read raw bytes to check for BOM
            with open(output_path, 'rb') as f:
                raw_bytes = f.read(3)

            # UTF-8 BOM is EF BB BF
            assert raw_bytes == b'\xef\xbb\xbf', f"Expected UTF-8 BOM (EF BB BF), got {raw_bytes.hex()}"

        finally:
            # Cleanup
            if output_path.exists():
                output_path.unlink()

    def test_csv_special_characters_preserved(self):
        """Test that special characters are preserved in CSV"""
        # Create test workbook with various special characters
        wb = Workbook()
        ws = wb.active
        ws.append(["Latin", "Greek", "Cyrillic"])
        ws.append(["Café Müller", "Ελληνικά", "Кириллица"])
        ws.append(["ñ ö ü é", "α β γ", "а б в"])

        # Save as CSV
        with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as f:
            output_path = Path(f.name)

        try:
            ExcelAnonymizer.save_workbook_as_csv(wb, output_path)

            # Read CSV with utf-8-sig encoding
            with open(output_path, 'r', encoding='utf-8-sig') as f:
                content = f.read()

            # Check all special characters are present
            special_chars = ['Café', 'Müller', 'ñ', 'ö', 'ü', 'é', 'Ελληνικά', 'Кириллица', 'α', 'β', 'γ', 'а', 'б', 'в']
            for char in special_chars:
                assert char in content, f"Special character '{char}' not found in CSV"

        finally:
            # Cleanup
            if output_path.exists():
                output_path.unlink()

    def test_csv_empty_cells_handled(self):
        """Test that empty cells are handled correctly"""
        wb = Workbook()
        ws = wb.active
        ws.append(["A", "B", "C"])
        ws.append(["1", None, "3"])
        ws.append([None, "2", None])

        # Save as CSV
        with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as f:
            output_path = Path(f.name)

        try:
            ExcelAnonymizer.save_workbook_as_csv(wb, output_path)

            # Read CSV
            with open(output_path, 'r', encoding='utf-8-sig') as f:
                lines = f.readlines()

            # Should have 3 lines (header + 2 data rows)
            assert len(lines) == 3

        finally:
            # Cleanup
            if output_path.exists():
                output_path.unlink()

    def test_csv_mixed_content_types(self):
        """Test CSV with mixed content types (text, numbers, dates)"""
        from datetime import datetime

        wb = Workbook()
        ws = wb.active
        ws.append(["Text", "Number", "Date"])
        ws.append(["Café", 123.45, datetime(2025, 1, 22)])
        ws.append(["Müller", 67.89, datetime(2025, 12, 31)])

        # Save as CSV
        with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as f:
            output_path = Path(f.name)

        try:
            ExcelAnonymizer.save_workbook_as_csv(wb, output_path)

            # Read CSV with utf-8-sig encoding
            with open(output_path, 'r', encoding='utf-8-sig') as f:
                content = f.read()

            # Check special characters preserved
            assert 'Café' in content
            assert 'Müller' in content
            # Check numbers are present
            assert '123.45' in content or '123,45' in content
            assert '67.89' in content or '67,89' in content

        finally:
            # Cleanup
            if output_path.exists():
                output_path.unlink()

    def test_csv_large_file_performance(self):
        """Test CSV export performance with large file"""
        import time

        # Create large workbook (1000 rows)
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Email", "Phone"])
        for i in range(1000):
            ws.append([f"User{i}", f"user{i}@example.com", f"06-1234567{i%10}"])

        # Save as CSV
        with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as f:
            output_path = Path(f.name)

        try:
            start = time.time()
            ExcelAnonymizer.save_workbook_as_csv(wb, output_path)
            duration = time.time() - start

            # Should complete within 1 second
            assert duration < 1.0

            # Verify file was created and has BOM
            with open(output_path, 'rb') as f:
                raw_bytes = f.read(3)
            assert raw_bytes == b'\xef\xbb\xbf'

        finally:
            # Cleanup
            if output_path.exists():
                output_path.unlink()
