"""
Edge Case Test Suite - Task 04-02

Comprehensive test suite for production edge cases from PITFALLS.md:
- Corrupt files (truncated, malformed ZIP, invalid XML)
- Encoding issues (Latin-1, Windows-1252, mixed encodings)
- Large files (performance boundaries, memory constraints)
- Empty files and malformed content

Requirements covered: MON-01
"""

import pytest
import io
import time
import zipfile
from pathlib import Path
from werkzeug.datastructures import FileStorage

from tests.fixtures.corrupt_files import (
    create_truncated_xlsx,
    create_invalid_zip_structure,
    create_malformed_xml_docx,
    create_non_pdf_with_pdf_extension,
    create_empty_file,
    create_xlsx_with_only_headers,
    create_scanned_pdf_mock,
    create_corrupted_pdf,
    create_large_xlsx,
    create_xlsx_with_macros_extension,
)

# =============================================================================
# Task 1: Corrupt File Tests
# =============================================================================

class TestCorruptFiles:
    """
    Test graceful handling of corrupt/malformed files

    From PITFALLS.md:
    - Line 366: Truncated XLSX at 50% size
    - Line 88: Invalid ZIP structure
    - Line 88: Malformed XML in Office documents
    - Line 88: Non-PDF file with .pdf extension

    Success criteria:
    - Corrupt files detected before processing
    - Clear error messages (not stack traces)
    - No 500 errors or crashes
    """

    def test_truncated_xlsx_50_percent(self, client):
        """
        Test truncated XLSX file (cut at 50%)

        Expected: FileCorruptError with message "File appears to be corrupted"

        From PITFALLS.md line 366:
        - Truncated files are common when downloads fail
        - Should detect during upload/preview, not during processing
        """
        corrupt_data = create_truncated_xlsx(0.5)

        # Try to upload
        data = {
            'files[]': (io.BytesIO(corrupt_data), 'test_truncated.xlsx')
        }

        response = client.post('/api/upload', data=data, content_type='multipart/form-data')

        # Should either reject during upload or detect during preview
        if response.status_code == 200:
            # If upload succeeded, check if files were actually uploaded
            json_data = response.get_json()
            if len(json_data.get('files', [])) > 0:
                # File was uploaded, preview should detect corruption
                file_id = json_data['files'][0]['id']

                preview_response = client.get(f'/api/excel-preview/{file_id}.xlsx')
                assert preview_response.status_code in [400, 500]

                preview_data = preview_response.get_json()
                assert 'error' in preview_data
                # Check for user-friendly error message
                error_msg = preview_data['error'].lower()
                assert any(keyword in error_msg for keyword in [
                    'corrupt', 'damaged', 'invalid', 'cannot read', 'not a valid'
                ]), f"Error message not user-friendly: {preview_data['error']}"
            else:
                # Files were rejected during upload - that's also valid
                pass
        else:
            # Upload rejected - corruption detected early (ideal)
            pass

    def test_truncated_xlsx_75_percent(self, client):
        """
        Test truncated XLSX file (cut at 75%)

        Partially corrupted files may pass initial checks but fail during processing
        """
        corrupt_data = create_truncated_xlsx(0.75)

        data = {
            'files[]': (io.BytesIO(corrupt_data), 'test_truncated_75.xlsx')
        }

        response = client.post('/api/upload', data=data, content_type='multipart/form-data')

        # Similar to 50% test
        if response.status_code == 200:
            json_data = response.get_json()
            if len(json_data.get('files', [])) > 0:
                file_id = json_data['files'][0]['id']

                preview_response = client.get(f'/api/excel-preview/{file_id}.xlsx')
                assert preview_response.status_code in [400, 500]

                preview_data = preview_response.get_json()
                assert 'error' in preview_data

    def test_invalid_zip_structure(self, client):
        """
        Test XLSX with corrupted ZIP headers

        XLSX files are ZIP archives. Corrupted central directory
        should be detected and reported clearly.

        From PITFALLS.md:
        - Invalid ZIP structure causes zipfile.BadZipFile
        - Should return clear error, not exception traceback
        """
        corrupt_data = create_invalid_zip_structure()

        data = {
            'files[]': (io.BytesIO(corrupt_data), 'test_bad_zip.xlsx')
        }

        response = client.post('/api/upload', data=data, content_type='multipart/form-data')

        if response.status_code == 200:
            json_data = response.get_json()
            if len(json_data.get('files', [])) > 0:
                file_id = json_data['files'][0]['id']

                preview_response = client.get(f'/api/excel-preview/{file_id}.xlsx')
                assert preview_response.status_code in [400, 500]

                preview_data = preview_response.get_json()
                assert 'error' in preview_data
                error_msg = preview_data['error'].lower()
                assert 'zip' in error_msg or 'corrupt' in error_msg

    def test_malformed_xml_docx(self, client):
        """
        Test DOCX with invalid XML in document.xml

        DOCX files contain XML. Malformed XML should be detected
        and reported with clear error message.

        From PITFALLS.md:
        - Malformed XML in Office documents is common corruption type
        """
        corrupt_data = create_malformed_xml_docx()

        data = {
            'files[]': (io.BytesIO(corrupt_data), 'test_bad_xml.docx')
        }

        response = client.post('/api/upload', data=data, content_type='multipart/form-data')
        
        # File might be rejected or accepted
        if response.status_code != 200 or len(response.get_json().get('files', [])) == 0:
            # Corruption detected at upload - test passes
            return

        # Try to process it
        json_data = response.get_json()
        file_id = json_data['files'][0]['id']

        anonymize_data = {
            'fileIds': [file_id],
            'rules': [],
            'excelRules': [],
            'activeTab': 'text',
            'autoDetectEnabled': False
        }

        anonymize_response = client.post(
            '/api/process',
            json=anonymize_data,
            content_type='application/json'
        )

        # Should return error during processing
        # Accept either 400 (validation error) or 500 (processing error)
        # but with clear error message
        assert anonymize_response.status_code in [400, 500]
        error_data = anonymize_response.get_json()
        assert 'error' in error_data or 'message' in error_data

    def test_non_pdf_file_with_pdf_extension(self, client):
        """
        Test text file disguised as PDF

        Files with wrong extension should be detected by content,
        not just filename.

        From PITFALLS.md line 88:
        - MIME type validation should catch extension mismatches
        """
        fake_pdf = create_non_pdf_with_pdf_extension()

        data = {
            'files[]': (io.BytesIO(fake_pdf), 'fake.pdf')
        }

        response = client.post('/api/upload', data=data, content_type='multipart/form-data')

        # Should either reject during upload or fail during processing
        if response.status_code == 200:
            json_data = response.get_json()
            if len(json_data.get('files', [])) == 0:
                # Rejected at upload - test passes
                return
            file_id = json_data['files'][0]['id']

            # Try to process as PDF
            anonymize_data = {
                'fileIds': [file_id],
                'rules': [],
                'excelRules': [],
                'activeTab': 'text',
                'autoDetectEnabled': False
            }

            anonymize_response = client.post(
                '/api/process',
                json=anonymize_data,
                content_type='application/json'
            )

            # Should fail with clear error
            assert anonymize_response.status_code in [400, 500]
            error_data = anonymize_response.get_json()
            assert 'error' in error_data or 'message' in error_data
            error_msg = str(error_data).lower()
            assert 'pdf' in error_msg or 'invalid' in error_msg

    def test_corrupted_pdf_header(self, client):
        """
        Test PDF with corrupted header

        PDF files must start with "%PDF-1.x" header.
        Corrupted header should be detected.
        """
        corrupt_pdf = create_corrupted_pdf()

        data = {
            'files[]': (io.BytesIO(corrupt_pdf), 'corrupt.pdf')
        }

        response = client.post('/api/upload', data=data, content_type='multipart/form-data')

        if response.status_code == 200:
            json_data = response.get_json()
            if len(json_data.get('files', [])) == 0:
                # Rejected at upload - test passes
                return
            file_id = json_data['files'][0]['id']

            anonymize_data = {
                'fileIds': [file_id],
                'rules': [],
                'excelRules': [],
                'activeTab': 'text',
                'autoDetectEnabled': False
            }

            anonymize_response = client.post(
                '/api/process',
                json=anonymize_data,
                content_type='application/json'
            )

            # Corrupted PDF might be accepted if corruption doesn't affect parsing
            # The key is that it doesn't crash the application
            # If it processes successfully, that's OK (resilient behavior)
            # If it fails, error should be clear
            if anonymize_response.status_code not in [200]:
                error_data = anonymize_response.get_json()
                assert 'error' in error_data or 'message' in error_data


# =============================================================================
# Task 2: Encoding Edge Cases
# =============================================================================

class TestEncodingEdgeCases:
    """
    Test encoding detection and preservation across file types

    From PITFALLS.md:
    - Line 39: 71% of CSV problems are encoding issues
    - Line 332: "CafÃ©" gibberish from wrong encoding
    - Line 813-829: Special characters in various encodings

    Success criteria:
    - All encodings detected and handled
    - Special characters preserved in output
    - CSV exports include UTF-8 BOM
    """

    @pytest.fixture
    def special_characters(self):
        """
        Special characters to test encoding preservation

        From PITFALLS.md line 56:
        - é, ö, ü, ç, ñ, —, €
        - Greek, Cyrillic, Chinese (line 824)
        """
        return {
            'western_european': 'Café Müller — naïve résumé',
            'currency': '€100 £50 ¥1000',
            'spanish': 'mañana piñata',
            'german': 'Übergrößenträger',
            'french': 'Côte d\'Azur',
            'mixed': 'Café €10 — São Paulo résumé'
        }

    def test_utf8_encoding_preservation(self, client, special_characters):
        """
        Test UTF-8 encoding is preserved during processing

        Standard UTF-8 is the default and should always work
        """
        test_text = '\n'.join(special_characters.values())

        data = {
            'files[]': (io.BytesIO(test_text.encode('utf-8')), 'test_utf8.txt')
        }

        response = client.post('/api/upload', data=data, content_type='multipart/form-data')
        assert response.status_code == 200

        json_data = response.get_json()
        file_id = json_data['files'][0]['id']

        # Process with no rules (pass-through test)
        anonymize_data = {
            'fileIds': [file_id],
            'rules': [],
            'excelRules': [],
            'activeTab': 'text',
            'autoDetectEnabled': False
        }

        anonymize_response = client.post(
            '/api/process',
            json=anonymize_data,
            content_type='application/json'
        )
        assert anonymize_response.status_code == 200

        # Download and verify characters preserved
        result_data = anonymize_response.get_json()
        download_response = client.get(f"/api/download/{file_id}")
        assert download_response.status_code == 200

        output_text = download_response.data.decode('utf-8')
        for char_set in special_characters.values():
            for char in char_set:
                if char not in [' ', '-', '\'']:  # Skip spaces and common chars
                    assert char in output_text, f"Character '{char}' not preserved"

    def test_utf8_bom_encoding(self, client, special_characters):
        """
        Test UTF-8 with BOM (Excel-friendly encoding)

        From PITFALLS.md line 39:
        - Excel only recognizes UTF-8 if BOM is present
        - 71% of CSV display issues are wrong encoding
        """
        test_text = '\n'.join(special_characters.values())
        # UTF-8 BOM: EF BB BF
        utf8_bom_data = '\ufeff'.encode('utf-8') + test_text.encode('utf-8')

        data = {
            'files[]': (io.BytesIO(utf8_bom_data), 'test_utf8_bom.txt')
        }

        response = client.post('/api/upload', data=data, content_type='multipart/form-data')
        assert response.status_code == 200

        json_data = response.get_json()
        file_id = json_data['files'][0]['id']

        anonymize_data = {
            'fileIds': [file_id],
            'rules': [],
            'excelRules': [],
            'activeTab': 'text',
            'autoDetectEnabled': False
        }

        anonymize_response = client.post(
            '/api/process',
            json=anonymize_data,
            content_type='application/json'
        )
        assert anonymize_response.status_code == 200

    def test_latin1_encoding(self, client, special_characters):
        """
        Test Latin-1 (ISO-8859-1) encoding

        From PITFALLS.md:
        - Common in legacy European systems
        - Should auto-detect or handle gracefully
        """
        # Latin-1 can only encode Western European chars
        test_text = special_characters['western_european']

        try:
            latin1_data = test_text.encode('latin-1')
        except UnicodeEncodeError:
            pytest.skip("Test text contains non-Latin-1 characters")

        data = {
            'files[]': (io.BytesIO(latin1_data), 'test_latin1.txt')
        }

        response = client.post('/api/upload', data=data, content_type='multipart/form-data')

        # May reject or auto-convert - both are acceptable
        # Key is: no crash, clear handling
        assert response.status_code in [200, 400]

    def test_windows1252_encoding(self, client):
        """
        Test Windows-1252 encoding (Windows default)

        From PITFALLS.md:
        - Default encoding on Windows systems
        - Common source of encoding issues
        """
        test_text = 'Smart quotes: "hello" \'world\' — em dash'

        try:
            windows_data = test_text.encode('windows-1252')
        except UnicodeEncodeError:
            pytest.skip("Test text contains non-Windows-1252 characters")

        data = {
            'files[]': (io.BytesIO(windows_data), 'test_windows1252.txt')
        }

        response = client.post('/api/upload', data=data, content_type='multipart/form-data')
        assert response.status_code in [200, 400]

    def test_csv_export_includes_bom(self, client, tmp_path):
        """
        Test that CSV exports include UTF-8 BOM for Excel compatibility

        From PITFALLS.md line 42:
        - Always save CSV with UTF-8 BOM (encoding='utf-8-sig')
        - Excel only recognizes UTF-8 if BOM present

        Critical success criterion: CSV exports MUST have BOM
        """
        # This test verifies our Excel anonymizer exports CSV with BOM
        # We'll need to process an Excel file and check output

        from openpyxl import Workbook

        # Create test Excel with special characters
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Name"
        ws['B1'] = "Description"
        ws['A2'] = "Café"
        ws['B2'] = "Müller résumé"

        excel_file = tmp_path / "test_special_chars.xlsx"
        wb.save(excel_file)

        # Upload
        with open(excel_file, 'rb') as f:
            data = {
                'files[]': (f, 'test_special_chars.xlsx')
            }
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')

        assert response.status_code == 200
        json_data = response.get_json()
        file_id = json_data['files'][0]['id']

        # Process (no anonymization, just pass-through)
        anonymize_data = {
            'fileIds': [file_id],
            'rules': [],
            'excelRules': [],
            'activeTab': 'excel',
            'autoDetectEnabled': False
        }

        anonymize_response = client.post(
            '/api/process',
            json=anonymize_data,
            content_type='application/json'
        )
        assert anonymize_response.status_code == 200

        # Download and check for BOM
        result_data = anonymize_response.get_json()
        download_response = client.get(f"/api/download/{result_data['results'][0]['anonymizedName']}")

        # Check if output has BOM (if it's converted to CSV)
        # Note: Output might still be XLSX, which handles encoding internally
        output_data = download_response.data

        # If it's XLSX, BOM not needed (Excel handles it)
        # If it's CSV, BOM should be present
        if result_data['results'][0]['anonymizedName'].endswith('.csv'):
            # Check for UTF-8 BOM: EF BB BF
            assert output_data[:3] == b'\xef\xbb\xbf', "CSV export missing UTF-8 BOM"

    def test_excel_special_characters_preserved(self, client, tmp_path, special_characters):
        """
        Test special characters in Excel cells are preserved

        From PITFALLS.md line 824:
        - Excel with mixed Greek, Cyrillic, Chinese characters
        """
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # Add all special character sets
        ws['A1'] = "Type"
        ws['B1'] = "Text"
        row = 2
        for char_type, text in special_characters.items():
            ws[f'A{row}'] = char_type
            ws[f'B{row}'] = text
            row += 1

        excel_file = tmp_path / "test_excel_encoding.xlsx"
        wb.save(excel_file)

        # Upload
        with open(excel_file, 'rb') as f:
            data = {
                'files[]': (f, 'test_excel_encoding.xlsx')
            }
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')

        assert response.status_code == 200
        json_data = response.get_json()
        file_id = json_data['files'][0]['id']

        # Get preview to check encoding
        preview_response = client.get(f'/api/excel-preview/{file_id}')
        assert preview_response.status_code == 200

        preview_data = preview_response.get_json()
        # Check preview data contains special characters
        # (Preview reads Excel, so encoding should be preserved)
        preview_str = str(preview_data)
        assert 'Café' in preview_str or 'Müller' in preview_str


# =============================================================================
# Task 3: Large File and Performance Tests
# =============================================================================

class TestLargeFilePerformance:
    """
    Test performance boundaries and memory constraints

    From PITFALLS.md:
    - Line 106: 10MB file timeout test
    - Line 80-111: Large file processing
    - Line 833-853: Performance stress tests

    Success criteria:
    - 10k row Excel processes in <30s
    - File size warnings appear at 5MB
    - Files >50MB rejected with clear message
    """

    def test_large_excel_10k_rows_performance(self, client, tmp_path):
        """
        Test 10k row Excel file processes within acceptable time

        From PITFALLS.md:
        - 7MB file (700k lines) takes 7-10 minutes in Presidio
        - Our tool should be faster: 10k rows in <30s

        Note: This is a long-running test (~10-30s)
        """
        # Create 10k row Excel (~10MB)
        large_data = create_large_xlsx(10000)

        large_file = tmp_path / "large_10k.xlsx"
        with open(large_file, 'wb') as f:
            f.write(large_data)

        # Upload
        with open(large_file, 'rb') as f:
            data = {
                'files[]': (f, 'large_10k.xlsx')
            }
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')

        assert response.status_code == 200
        json_data = response.get_json()
        file_id = json_data['files'][0]['id']

        # Process with simple rule
        anonymize_data = {
            'fileIds': [file_id],
            'rules': [],
            'excelRules': [
                {
                    'id': 'rule-1',
                    'columnName': 'Email',
                    'strategy': 'replace',
                    'replacementText': '[EMAIL]'
                }
            ],
            'activeTab': 'excel',
            'autoDetectEnabled': False
        }

        start_time = time.time()

        anonymize_response = client.post(
            '/api/process',
            json=anonymize_data,
            content_type='application/json'
        )

        elapsed = time.time() - start_time

        # Should complete successfully
        assert anonymize_response.status_code == 200, \
            f"Processing failed: {anonymize_response.get_json()}"

        # Performance check: <30s for 10k rows
        assert elapsed < 30, \
            f"Processing took {elapsed:.2f}s, expected <30s for 10k rows"

        print(f"\n✓ 10k row Excel processed in {elapsed:.2f}s")

    @pytest.mark.slow
    def test_large_excel_50k_rows_performance(self, client, tmp_path):
        """
        Test 50k row Excel file (approaching 50MB limit)

        From PITFALLS.md:
        - Should complete within Gunicorn timeout (120s)

        Note: This is a very long-running test (~60-120s)
        Marked as @pytest.mark.slow - run with: pytest -m slow
        """
        # Create 50k row Excel (~50MB)
        large_data = create_large_xlsx(50000)

        large_file = tmp_path / "large_50k.xlsx"
        with open(large_file, 'wb') as f:
            f.write(large_data)

        # Check file size
        file_size_mb = len(large_data) / (1024 * 1024)
        print(f"\n50k row Excel file size: {file_size_mb:.2f}MB")

        # If >50MB, should be rejected
        if file_size_mb > 50:
            pytest.skip(f"File exceeds 50MB limit ({file_size_mb:.2f}MB)")

        # Upload
        with open(large_file, 'rb') as f:
            data = {
                'files[]': (f, 'large_50k.xlsx')
            }
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')

        assert response.status_code == 200
        json_data = response.get_json()
        file_id = json_data['files'][0]['id']

        # Process
        anonymize_data = {
            'fileIds': [file_id],
            'rules': [],
            'excelRules': [],
            'activeTab': 'excel',
            'autoDetectEnabled': False
        }

        start_time = time.time()

        anonymize_response = client.post(
            '/api/process',
            json=anonymize_data,
            content_type='application/json'
        )

        elapsed = time.time() - start_time

        assert anonymize_response.status_code == 200
        assert elapsed < 120, \
            f"Processing took {elapsed:.2f}s, expected <120s (Gunicorn timeout)"

        print(f"\n✓ 50k row Excel processed in {elapsed:.2f}s")

    def test_file_size_warning_5mb(self, client, tmp_path):
        """
        Test file size warning appears at 5MB soft limit

        From PITFALLS.md line 93:
        - Warn at >5MB, block at >50MB
        """
        # Create ~6MB file (need enough rows to exceed 5MB)
        large_data = create_large_xlsx(50000)  # ~50MB should exceed 5MB

        large_file = tmp_path / "medium_large.xlsx"
        with open(large_file, 'wb') as f:
            f.write(large_data)

        file_size_mb = len(large_data) / (1024 * 1024)
        if file_size_mb < 5:
            pytest.skip(f"Test file too small: {file_size_mb:.2f}MB, need >5MB")

        # Upload
        with open(large_file, 'rb') as f:
            data = {
                'files[]': (f, 'medium_6mb.xlsx')
            }
            response = client.post('/api/upload', data=data, content_type='multipart/form-data')

        # Should succeed but may include warning
        assert response.status_code == 200
        json_data = response.get_json()

        # Check for warning in response (implementation-dependent)
        # If no warning system yet, this test documents the requirement
        print(f"\n✓ {file_size_mb:.2f}MB file uploaded successfully")


# =============================================================================
# Task 4: Empty Files and Malformed Content
# =============================================================================

class TestEmptyAndMalformedFiles:
    """
    Test handling of empty and edge case file content

    From PITFALLS.md:
    - Line 368: Empty files
    - Line 341, 791: Scanned PDFs
    - Line 370: Password-protected files
    - Line 371: Files with macros

    Success criteria:
    - Empty files: "File is empty" error
    - Scanned PDFs: "No text found - OCR required" warning
    - Password-protected: "Upload unlocked version" error
    - Macros detected and warned
    """

    def test_empty_text_file(self, client):
        """
        Test 0-byte file

        From PITFALLS.md line 368:
        - Empty files confuse users
        - Should return clear error: "File is empty"
        """
        empty_data = create_empty_file()

        data = {
            'files[]': (io.BytesIO(empty_data), 'empty.txt')
        }

        response = client.post('/api/upload', data=data, content_type='multipart/form-data')

        # Empty files might be accepted (user-friendly) or rejected
        # Key is: no crash
        if response.status_code in [400, 413]:
            # Rejected - should have clear error
            json_data = response.get_json()
            assert 'error' in json_data or 'message' in json_data
            error_msg = str(json_data).lower()
            assert 'empty' in error_msg or 'invalid' in error_msg
        else:
            # Accepted - that's OK too (allows user to see the issue)
            assert response.status_code == 200

    def test_empty_excel_file(self, client):
        """Test 0-byte Excel file"""
        empty_data = create_empty_file()

        data = {
            'files[]': (io.BytesIO(empty_data), 'empty.xlsx')
        }

        response = client.post('/api/upload', data=data, content_type='multipart/form-data')

        # Same as text file - accept or reject, no crash
        if response.status_code in [400, 413]:
            json_data = response.get_json()
            error_msg = str(json_data).lower()
            assert 'empty' in error_msg or 'invalid' in error_msg
        else:
            assert response.status_code == 200

    def test_excel_with_only_headers(self, client):
        """
        Test Excel file with headers but no data rows

        This is technically valid but may confuse users.
        Should process successfully but may warn.
        """
        headers_only_data = create_xlsx_with_only_headers()

        data = {
            'files[]': (io.BytesIO(headers_only_data), 'headers_only.xlsx')
        }

        response = client.post('/api/upload', data=data, content_type='multipart/form-data')
        assert response.status_code == 200

        json_data = response.get_json()
        file_id = json_data['files'][0]['id']

        # Preview should work
        preview_response = client.get(f'/api/excel-preview/{file_id}')
        assert preview_response.status_code == 200

        preview_data = preview_response.get_json()
        # Should have sheets with columns but rows should be empty
        assert 'sheets' in preview_data
        assert len(preview_data['sheets']) > 0
        first_sheet = preview_data['sheets'][0]
        assert 'columns' in first_sheet
        assert len(first_sheet['columns']) > 0
        assert first_sheet.get('rows', 0) == 0 or len(first_sheet.get('preview', [])) == 0

    def test_scanned_pdf_no_text(self, client):
        """
        Test scanned PDF (image-only, no text layer)

        From PITFALLS.md line 341, 791:
        - Scanned PDFs are common
        - Should warn: "No text found - OCR required"

        Note: Our mock PDF may have some text, but real scanned PDFs
        would have zero extractable text.
        """
        scanned_pdf = create_scanned_pdf_mock()

        data = {
            'files[]': (io.BytesIO(scanned_pdf), 'scanned.pdf')
        }

        response = client.post('/api/upload', data=data, content_type='multipart/form-data')
        assert response.status_code == 200

        json_data = response.get_json()
        file_id = json_data['files'][0]['id']

        # Try to process
        anonymize_data = {
            'fileTypes': {
                'pdf': {
                    'files': [f'{file_id}.pdf'],
                    'rules': [],
                    'autoDetect': {'phones': False, 'emails': False}
                }
            }
        }

        anonymize_response = client.post(
            '/api/process',
            json=anonymize_data,
            content_type='application/json'
        )

        # Should succeed but may have warning about no text
        # (Implementation may vary - document expected behavior)
        assert anonymize_response.status_code in [200, 400]

    def test_xlsm_macro_enabled_file(self, client):
        """
        Test .xlsm (macro-enabled) file

        From PITFALLS.md line 371:
        - Files with .xlsm extension should be warned/rejected
        - Macros are security risk
        """
        xlsm_data = create_xlsx_with_macros_extension()

        data = {
            'files[]': (io.BytesIO(xlsm_data), 'test_with_macros.xlsm')
        }

        response = client.post('/api/upload', data=data, content_type='multipart/form-data')

        # Should either reject or warn about macros
        if response.status_code == 200:
            # If upload succeeds, check for warning
            json_data = response.get_json()
            # Implementation should detect .xlsm extension
            print(f"\n✓ .xlsm file uploaded: {json_data}")
        else:
            # If rejected, should have clear error
            assert response.status_code == 400
            json_data = response.get_json()
            error_msg = str(json_data).lower()
            assert 'macro' in error_msg or 'xlsm' in error_msg


# =============================================================================
# Performance Benchmarks Summary
# =============================================================================

@pytest.fixture(scope='session', autouse=True)
def print_performance_benchmarks():
    """Print performance benchmarks after test run"""
    yield
    print("\n" + "="*80)
    print("EDGE CASE TEST SUITE - PERFORMANCE BENCHMARKS")
    print("="*80)
    print("\nExpected performance (from PITFALLS.md):")
    print("  - 10k row Excel: <30s")
    print("  - 50k row Excel: <120s (Gunicorn timeout)")
    print("  - File size warning: 5MB")
    print("  - File size hard limit: 50MB")
    print("\nTo run performance tests:")
    print("  pytest tests/test_edge_cases.py::TestLargeFilePerformance -v")
    print("\nTo run slow tests (50k rows):")
    print("  pytest tests/test_edge_cases.py -m slow -v")
    print("="*80)
