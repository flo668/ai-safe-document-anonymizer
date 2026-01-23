"""
Tests voor Excel Anonymizer - Shuffle & Type Validation

Test Coverage:
- Shuffle transformation (XLS-01, XLS-02)
- Type detection (number, date, text)
- Type validation (XLS-09)
- Statistical preservation (mean, stddev)
"""

import pytest
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from anonymizer.excel_anonymizer import (
    ExcelAnonymizer,
    ExcelColumnRule,
)


# ============================================================================
# Fixtures
# ============================================================================

@pytest.fixture
def sample_workbook_with_numbers(tmp_path):
    """Create Excel file with numeric column"""
    wb = Workbook()
    ws = wb.active

    # Headers
    ws.append(['ID', 'Price', 'Quantity'])

    # Data rows
    data = [
        [1, 10.5, 100],
        [2, 20.0, 200],
        [3, 15.5, 150],
        [4, 25.0, 250],
        [5, 30.5, 300],
        [6, 12.5, 120],
        [7, 18.0, 180],
        [8, 22.5, 225],
    ]
    for row in data:
        ws.append(row)

    filepath = tmp_path / "numbers.xlsx"
    wb.save(filepath)
    return filepath


@pytest.fixture
def sample_workbook_with_dates(tmp_path):
    """Create Excel file with date column"""
    wb = Workbook()
    ws = wb.active

    # Headers
    ws.append(['ID', 'OrderDate', 'ShipDate'])

    # Data rows with datetime objects
    base_date = datetime(2025, 1, 1)
    for i in range(8):
        ws.append([
            i + 1,
            base_date + timedelta(days=i),
            base_date + timedelta(days=i + 7)
        ])

    filepath = tmp_path / "dates.xlsx"
    wb.save(filepath)
    return filepath


@pytest.fixture
def sample_workbook_with_text(tmp_path):
    """Create Excel file with text column"""
    wb = Workbook()
    ws = wb.active

    # Headers
    ws.append(['ID', 'Name', 'Email'])

    # Data rows
    data = [
        [1, 'Alice', 'alice@example.com'],
        [2, 'Bob', 'bob@example.com'],
        [3, 'Charlie', 'charlie@example.com'],
        [4, 'David', 'david@example.com'],
        [5, 'Eve', 'eve@example.com'],
    ]
    for row in data:
        ws.append(row)

    filepath = tmp_path / "text.xlsx"
    wb.save(filepath)
    return filepath


@pytest.fixture
def sample_workbook_mixed(tmp_path):
    """Create Excel file with mixed column types"""
    wb = Workbook()
    ws = wb.active

    # Headers
    ws.append(['ID', 'Name', 'Price', 'OrderDate'])

    # Data rows
    base_date = datetime(2025, 1, 1)
    data = [
        [1, 'Product A', 100.0, base_date],
        [2, 'Product B', 200.0, base_date + timedelta(days=1)],
        [3, 'Product C', 150.0, base_date + timedelta(days=2)],
        [4, 'Product D', 250.0, base_date + timedelta(days=3)],
        [5, 'Product E', 300.0, base_date + timedelta(days=4)],
    ]
    for row in data:
        ws.append(row)

    filepath = tmp_path / "mixed.xlsx"
    wb.save(filepath)
    return filepath


# ============================================================================
# Shuffle Tests (XLS-01, XLS-02)
# ============================================================================

def test_shuffle_column_basic():
    """Test basic shuffle functionality"""
    values = [1, 2, 3, 4, 5]
    shuffled = ExcelAnonymizer.shuffle_column(values)

    # Should return list
    assert isinstance(shuffled, list)

    # Should have same length
    assert len(shuffled) == len(values)

    # Should contain same values
    assert sorted(shuffled) == sorted(values)

    # Should be different order (probabilistically - may fail 1/120 times)
    # We skip this check since it's not deterministic
    # assert shuffled != values


def test_shuffle_preserves_statistical_distribution():
    """Test that shuffle preserves mean and stddev (XLS-02)"""
    import statistics

    # Generate larger sample for statistical testing
    values = [10.5, 20.0, 15.5, 25.0, 30.5, 12.5, 18.0, 22.5, 27.0, 35.5]

    # Calculate original stats
    original_mean = statistics.mean(values)
    original_stddev = statistics.stdev(values)
    original_min = min(values)
    original_max = max(values)

    # Shuffle
    shuffled = ExcelAnonymizer.shuffle_column(values)

    # Calculate shuffled stats
    shuffled_mean = statistics.mean(shuffled)
    shuffled_stddev = statistics.stdev(shuffled)
    shuffled_min = min(shuffled)
    shuffled_max = max(shuffled)

    # Mean unchanged (within 0.01 tolerance)
    assert abs(shuffled_mean - original_mean) < 0.01

    # Stddev unchanged (within 0.01 tolerance)
    assert abs(shuffled_stddev - original_stddev) < 0.01

    # Min/max unchanged
    assert shuffled_min == original_min
    assert shuffled_max == original_max


def test_shuffle_in_excel_file(sample_workbook_with_numbers, tmp_path):
    """Test shuffle transformation in Excel file (XLS-01)"""
    import statistics
    from openpyxl import load_workbook

    # Load original to get stats
    wb_original = load_workbook(sample_workbook_with_numbers)
    ws_original = wb_original.active
    original_prices = [ws_original.cell(row=i, column=2).value for i in range(2, 10)]
    original_mean = statistics.mean(original_prices)
    original_stddev = statistics.stdev(original_prices)

    # Create rule
    rule = ExcelColumnRule({
        'id': 'shuffle-1',
        'columnName': 'Price',
        'anonymizationType': 'shuffle'
    })

    # Process
    output_path = tmp_path / "output.xlsx"
    ExcelAnonymizer.process_excel_file(
        sample_workbook_with_numbers,
        output_path,
        [rule],
        preserve_headers=True
    )

    # Load output
    wb_output = load_workbook(output_path)
    ws_output = wb_output.active
    shuffled_prices = [ws_output.cell(row=i, column=2).value for i in range(2, 10)]
    shuffled_mean = statistics.mean(shuffled_prices)
    shuffled_stddev = statistics.stdev(shuffled_prices)

    # Verify statistical preservation
    assert abs(shuffled_mean - original_mean) < 0.01
    assert abs(shuffled_stddev - original_stddev) < 0.01

    # Verify values are reordered (same values, different positions)
    assert sorted(shuffled_prices) == sorted(original_prices)


def test_shuffle_does_not_create_new_values():
    """Test that shuffle only reorders, doesn't create new values"""
    values = ['A', 'B', 'C', 'D', 'E']
    shuffled = ExcelAnonymizer.shuffle_column(values)

    # Every value in shuffled must be in original
    for val in shuffled:
        assert val in values

    # Every value in original must be in shuffled
    for val in values:
        assert val in shuffled


# ============================================================================
# Type Detection Tests
# ============================================================================

def test_detect_column_type_numbers():
    """Test type detection for numeric column"""
    values = [10, 20, 30, 40, 50, 60, 70, 80]
    detected = ExcelAnonymizer.detect_column_type(values)
    assert detected == 'number'


def test_detect_column_type_floats():
    """Test type detection for float column"""
    values = [10.5, 20.0, 15.5, 25.0, 30.5]
    detected = ExcelAnonymizer.detect_column_type(values)
    assert detected == 'number'


def test_detect_column_type_dates():
    """Test type detection for date column"""
    base = datetime(2025, 1, 1)
    values = [base + timedelta(days=i) for i in range(10)]
    detected = ExcelAnonymizer.detect_column_type(values)
    assert detected == 'date'


def test_detect_column_type_text():
    """Test type detection for text column"""
    values = ['Alice', 'Bob', 'Charlie', 'David', 'Eve']
    detected = ExcelAnonymizer.detect_column_type(values)
    assert detected == 'text'


def test_detect_column_type_empty():
    """Test type detection for empty column"""
    values = [None, '', ' ', None]
    detected = ExcelAnonymizer.detect_column_type(values)
    assert detected == 'text'


def test_detect_column_type_mixed_with_majority_numbers():
    """Test type detection with >70% numbers"""
    # 8 numbers, 2 text = 80% numbers
    values = [10, 20, 30, 40, 50, 60, 70, 80, 'text', 'text']
    detected = ExcelAnonymizer.detect_column_type(values)
    assert detected == 'number'


def test_detect_column_type_mixed_with_minority_numbers():
    """Test type detection with <70% numbers"""
    # 5 numbers, 5 text = 50% numbers
    values = [10, 20, 30, 40, 50, 'a', 'b', 'c', 'd', 'e']
    detected = ExcelAnonymizer.detect_column_type(values)
    assert detected == 'text'


# ============================================================================
# Type Validation Tests (XLS-09)
# ============================================================================

def test_validate_number_multiply_on_number_column():
    """Test number_multiply is valid on number column"""
    rule = ExcelColumnRule({
        'id': 'test-1',
        'columnName': 'Price',
        'columnType': 'number',
        'anonymizationType': 'number_multiply',
        'numberMultiplier': 1.5
    })
    values = [10, 20, 30, 40, 50]

    is_valid, error_msg = ExcelAnonymizer.validate_rule_type_compatibility(rule, values)

    assert is_valid is True
    assert error_msg == ""


def test_validate_number_multiply_on_text_column():
    """Test number_multiply fails on text column (XLS-09)"""
    rule = ExcelColumnRule({
        'id': 'test-2',
        'columnName': 'Name',
        'columnType': 'text',
        'anonymizationType': 'number_multiply',
        'numberMultiplier': 1.5
    })
    values = ['Alice', 'Bob', 'Charlie', 'David']

    is_valid, error_msg = ExcelAnonymizer.validate_rule_type_compatibility(rule, values)

    assert is_valid is False
    assert "Type mismatch" in error_msg
    assert "number_multiply" in error_msg
    assert "text" in error_msg


def test_validate_date_offset_on_date_column():
    """Test date_offset is valid on date column"""
    rule = ExcelColumnRule({
        'id': 'test-3',
        'columnName': 'OrderDate',
        'columnType': 'date',
        'anonymizationType': 'date_offset',
        'dateOffsetDays': 30
    })
    base = datetime(2025, 1, 1)
    values = [base + timedelta(days=i) for i in range(5)]

    is_valid, error_msg = ExcelAnonymizer.validate_rule_type_compatibility(rule, values)

    assert is_valid is True
    assert error_msg == ""


def test_validate_date_offset_on_number_column():
    """Test date_offset fails on number column (XLS-09)"""
    rule = ExcelColumnRule({
        'id': 'test-4',
        'columnName': 'Price',
        'columnType': 'number',
        'anonymizationType': 'date_offset',
        'dateOffsetDays': 30
    })
    values = [10, 20, 30, 40, 50]

    is_valid, error_msg = ExcelAnonymizer.validate_rule_type_compatibility(rule, values)

    assert is_valid is False
    assert "Type mismatch" in error_msg
    assert "date_offset" in error_msg
    assert "number" in error_msg


def test_validate_text_transformations_on_any_type():
    """Test text transformations (replace, jabber) work on any type"""
    text_types = ['replace', 'jabber', 'mask_start', 'mask_end', 'hash']

    for anon_type in text_types:
        rule = ExcelColumnRule({
            'id': f'test-{anon_type}',
            'columnName': 'TestCol',
            'anonymizationType': anon_type
        })

        # Test on number column
        is_valid, _ = ExcelAnonymizer.validate_rule_type_compatibility(rule, [10, 20, 30])
        assert is_valid is True

        # Test on text column
        is_valid, _ = ExcelAnonymizer.validate_rule_type_compatibility(rule, ['a', 'b', 'c'])
        assert is_valid is True

        # Test on date column
        base = datetime(2025, 1, 1)
        dates = [base + timedelta(days=i) for i in range(3)]
        is_valid, _ = ExcelAnonymizer.validate_rule_type_compatibility(rule, dates)
        assert is_valid is True


def test_validate_shuffle_on_any_type():
    """Test shuffle works on any column type"""
    rule = ExcelColumnRule({
        'id': 'test-shuffle',
        'columnName': 'TestCol',
        'anonymizationType': 'shuffle'
    })

    # Test on number column
    is_valid, _ = ExcelAnonymizer.validate_rule_type_compatibility(rule, [10, 20, 30])
    assert is_valid is True

    # Test on text column
    is_valid, _ = ExcelAnonymizer.validate_rule_type_compatibility(rule, ['a', 'b', 'c'])
    assert is_valid is True

    # Test on date column
    base = datetime(2025, 1, 1)
    dates = [base + timedelta(days=i) for i in range(3)]
    is_valid, _ = ExcelAnonymizer.validate_rule_type_compatibility(rule, dates)
    assert is_valid is True


# ============================================================================
# Integration Tests - Type Validation in Processing
# ============================================================================

def test_process_skips_invalid_rule(sample_workbook_with_text, tmp_path, capsys):
    """Test that processing skips invalid rule and logs error"""
    # Create invalid rule (number_multiply on text column)
    rule = ExcelColumnRule({
        'id': 'invalid-1',
        'columnName': 'Name',
        'columnType': 'text',
        'anonymizationType': 'number_multiply',
        'numberMultiplier': 1.5
    })

    # Process
    output_path = tmp_path / "output.xlsx"
    ExcelAnonymizer.process_excel_file(
        sample_workbook_with_text,
        output_path,
        [rule],
        preserve_headers=True
    )

    # Check that validation error was logged
    captured = capsys.readouterr()
    assert "[VALIDATION ERROR]" in captured.out
    assert "Type mismatch" in captured.out

    # Verify output file exists and data unchanged
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb.active

    # Name column should be unchanged
    assert ws.cell(row=2, column=2).value == 'Alice'
    assert ws.cell(row=3, column=2).value == 'Bob'


def test_process_applies_valid_rules_only(sample_workbook_mixed, tmp_path, capsys):
    """Test that processing applies valid rules and skips invalid ones"""
    # Create mix of valid and invalid rules
    rules = [
        ExcelColumnRule({
            'id': 'valid-1',
            'columnName': 'Name',
            'anonymizationType': 'replace',
            'replaceWith': '[NAME]'
        }),
        ExcelColumnRule({
            'id': 'invalid-1',
            'columnName': 'Name',
            'anonymizationType': 'number_multiply',  # Invalid: text column
            'numberMultiplier': 1.5
        }),
        ExcelColumnRule({
            'id': 'valid-2',
            'columnName': 'Price',
            'columnType': 'number',  # Specify number type
            'anonymizationType': 'number_multiply',  # Valid: number column
            'numberMultiplier': 2.0
        })
    ]

    # Process
    output_path = tmp_path / "output.xlsx"
    ExcelAnonymizer.process_excel_file(
        sample_workbook_mixed,
        output_path,
        rules,
        preserve_headers=True
    )

    # Check validation error was logged
    captured = capsys.readouterr()
    print(f"DEBUG: Captured output:\n{captured.out}")
    assert "[VALIDATION ERROR]" in captured.out

    # Verify valid rules were applied
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb.active

    # Name should be replaced (valid rule)
    print(f"DEBUG: Name cell value: {ws.cell(row=2, column=2).value}")
    assert ws.cell(row=2, column=2).value == '[NAME]'

    # Price should be multiplied (valid rule)
    print(f"DEBUG: Price cell value: {ws.cell(row=2, column=3).value}, type: {type(ws.cell(row=2, column=3).value)}")
    assert ws.cell(row=2, column=3).value == 200.0  # 100 * 2


# ============================================================================
# Edge Cases
# ============================================================================

def test_shuffle_empty_list():
    """Test shuffle with empty list"""
    values = []
    shuffled = ExcelAnonymizer.shuffle_column(values)
    assert shuffled == []


def test_shuffle_single_element():
    """Test shuffle with single element"""
    values = [42]
    shuffled = ExcelAnonymizer.shuffle_column(values)
    assert shuffled == [42]


def test_detect_column_type_with_none_values():
    """Test type detection ignores None values"""
    values = [None, 10, None, 20, None, 30, None]
    detected = ExcelAnonymizer.detect_column_type(values)
    assert detected == 'number'


def test_validate_rule_with_empty_column():
    """Test validation with empty column defaults to text"""
    rule = ExcelColumnRule({
        'id': 'test-empty',
        'columnName': 'Empty',
        'anonymizationType': 'number_multiply',
        'numberMultiplier': 1.5
    })
    values = [None, '', None, '']

    is_valid, error_msg = ExcelAnonymizer.validate_rule_type_compatibility(rule, values)

    # Empty column detected as text, so number_multiply is invalid
    assert is_valid is False
    assert "Type mismatch" in error_msg
