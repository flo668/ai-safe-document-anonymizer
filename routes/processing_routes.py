"""
Processing Routes - File Processing en Preview

Blueprint voor file processing, auto-detection preview, en Excel preview endpoints.
"""

from flask import Blueprint, request, jsonify, session, current_app
from anonymizer.reverse_anonymizer import AnonymizationMapping
from utils.encryption import SecureMappingStorage
from utils.audit import AuditLogger
from utils.metrics import get_metrics_collector, MemoryProfiler
import time

processing_bp = Blueprint('processing', __name__)


@processing_bp.route('/preview', methods=['POST'])
def preview_detection():
    """
    Preview endpoint - laat zien wat automatisch gedetecteerd zou worden
    Accepts: JSON met fileIds[], phonePlaceholder, emailPlaceholder
    Returns: JSON met preview info per bestand
    """
    from app import get_file_preview

    data = request.get_json()
    if not data:
        return jsonify({'error': 'Geen data ontvangen'}), 400

    session_id = session.get('session_id')
    if not session_id:
        return jsonify({'error': 'Geen actieve sessie'}), 400

    file_ids = data.get('fileIds', [])
    phone_placeholder = data.get('phonePlaceholder', '[TEL VERWIJDERD]')
    email_placeholder = data.get('emailPlaceholder', '[EMAIL VERWIJDERD]')

    if not file_ids:
        return jsonify({'error': 'Geen bestanden om te previewed'}), 400

    session_dir = current_app.config['UPLOAD_FOLDER'] / session_id
    previews = []

    for file_id in file_ids:
        try:
            preview_info = get_file_preview(
                session_dir, file_id, phone_placeholder, email_placeholder
            )
            previews.append(preview_info)
        except Exception as e:
            previews.append({
                'id': file_id,
                'error': str(e)
            })

    return jsonify({
        'success': True,
        'previews': previews
    })


@processing_bp.route('/process', methods=['POST'])
def process_files():
    """
    Processing endpoint
    Accepts: JSON met fileIds[], rules[], excelRules[], activeTab,
             phonePlaceholder, emailPlaceholder, generalPlaceholder, autoDetectEnabled
    Returns: JSON met processing results en logs
    """
    from app import process_single_file
    import json

    data = request.get_json()
    if not data:
        return jsonify({'error': 'Geen data ontvangen'}), 400

    session_id = session.get('session_id')
    if not session_id:
        return jsonify({'error': 'Geen actieve sessie'}), 400

    file_ids = data.get('fileIds', [])
    rules_data = data.get('rules', [])
    excel_rules_data = data.get('excelRules', [])
    active_tab = data.get('activeTab', 'text')

    # Nieuwe parameters voor auto-detectie
    phone_placeholder = data.get('phonePlaceholder', '[TEL VERWIJDERD]')
    email_placeholder = data.get('emailPlaceholder', '[EMAIL VERWIJDERD]')
    general_placeholder = data.get('generalPlaceholder', '[ANONIEM]')
    auto_detect_enabled = data.get('autoDetectEnabled', True)
    reversible_mode = data.get('reversibleMode', False)
    preserve_formatting = data.get('preserveFormatting', False)

    # Multi-sheet selectie (dict met filename → [sheet_names])
    selected_sheets = data.get('selectedSheets', {})

    if not file_ids:
        return jsonify({'error': 'Geen bestanden om te verwerken'}), 400

    session_dir = current_app.config['UPLOAD_FOLDER'] / session_id
    output_dir = current_app.config['OUTPUT_FOLDER'] / session_id
    output_dir.mkdir(exist_ok=True)

    # Initialiseer audit logger
    audit_logger = AuditLogger(session_id, output_dir)

    # Initialiseer metrics collector
    metrics = get_metrics_collector()

    # Maak mapping object voor reversible mode
    global_mapping = None
    if reversible_mode:
        global_mapping = AnonymizationMapping()

    results = []
    all_logs = []
    auto_detect_stats = {
        'total_phones': 0,
        'total_emails': 0,
        'total_manual_replacements': 0
    }

    for file_id in file_ids:
        # Start timing
        start_time = time.time()
        file_size = 0
        file_ext = active_tab  # Default to active tab

        try:
            # Log start van anonymization
            audit_logger.log_anonymize_start(file_id, active_tab)

            # Get file size for metrics
            input_file = None
            for filepath in session_dir.glob(f"{file_id}.*"):
                if filepath.is_file() and filepath.suffix not in ['.meta.json']:
                    input_file = filepath
                    file_size = filepath.stat().st_size
                    file_ext = filepath.suffix.lstrip('.').lower()
                    break

            # Get selected sheets for this file (if applicable)
            file_sheet_names = selected_sheets.get(file_id, None)  # None = all sheets

            # Use memory profiling for large files (>5MB) (MON-10)
            if file_size > 5 * 1024 * 1024:  # 5MB threshold
                with MemoryProfiler(f'process_file_{file_ext}_{file_size // (1024*1024)}MB'):
                    result = process_single_file(
                        session_dir, output_dir, file_id,
                        rules_data, excel_rules_data, active_tab,
                        phone_placeholder, email_placeholder, general_placeholder,
                        auto_detect_enabled, reversible_mode, global_mapping,
                        preserve_formatting, file_sheet_names
                    )
            else:
                result = process_single_file(
                    session_dir, output_dir, file_id,
                    rules_data, excel_rules_data, active_tab,
                    phone_placeholder, email_placeholder, general_placeholder,
                    auto_detect_enabled, reversible_mode, global_mapping,
                    preserve_formatting, file_sheet_names
                )

            # Calculate duration
            duration = time.time() - start_time

            results.append(result)
            all_logs.extend(result.get('logs', []))

            # Log completion van anonymization
            num_rules = len(rules_data) + len(excel_rules_data)
            audit_logger.log_anonymize_complete(
                result.get('outputFile', file_id),
                rules_applied=num_rules
            )

            # Log metrics (MON-06, MON-08)
            entities_found = sum(log.get('count', 0) for log in result.get('logs', []))
            metrics.log_processing_metrics(
                file_type=file_ext,
                file_size=file_size,
                duration=duration,
                entities_found=entities_found,
                success=True
            )

            # Verzamel statistieken
            if result.get('auto_detect_report'):
                report = result['auto_detect_report']
                auto_detect_stats['total_phones'] += report.get('phone_numbers', {}).get('count', 0)
                auto_detect_stats['total_emails'] += report.get('emails', {}).get('count', 0)

        except Exception as e:
            # Calculate duration for error case
            duration = time.time() - start_time

            # Log metrics for failure (MON-06, MON-07)
            metrics.log_processing_metrics(
                file_type=file_ext,
                file_size=file_size,
                duration=duration,
                entities_found=0,
                success=False,
                error=str(e)
            )

            # Check error rate and trigger alert if needed (MON-07)
            error_rate = metrics.get_error_rate(file_ext, window_minutes=60)
            if error_rate > metrics.ERROR_RATE_THRESHOLD:
                metrics.trigger_alert(
                    f"High error rate for {file_ext}",
                    error_rate,
                    file_type=file_ext
                )
            # Log error
            audit_logger.log_anonymize_error(file_id, str(e))
            # Probeer originalName op te halen voor error display
            metadata_file = session_dir / f"{file_id}.meta.json"
            original_name = 'Unknown'
            if metadata_file.exists():
                try:
                    with open(metadata_file, 'r', encoding='utf-8') as f:
                        metadata = json.load(f)
                        original_name = metadata.get('originalName', 'Unknown')
                except:
                    pass

            print(f"ERROR processing {file_id} ({original_name}): {str(e)}")
            results.append({
                'id': file_id,
                'status': 'error',
                'error': str(e),
                'originalName': original_name
            })

    # Tel handmatige vervangingen
    for log in all_logs:
        if log.get('ruleId') not in ['auto_phone', 'auto_email']:
            auto_detect_stats['total_manual_replacements'] += log.get('count', 0)

    response_data = {
        'success': True,
        'results': results,
        'logs': all_logs,
        'statistics': auto_detect_stats
    }

    # Voeg mapping toe indien reversible mode
    if reversible_mode and global_mapping:
        try:
            # Gebruik encryption voor mapping storage
            secure_storage = SecureMappingStorage(session_id)
            mapping_dict = global_mapping.to_dict()

            # Sla encrypted mapping op
            secure_storage.save_mapping(
                mapping_dict,
                output_dir,
                filename=f"mapping_{session_id}.json"
            )

            # Log mapping save
            audit_logger.log_mapping_saved(
                len(global_mapping.mappings),
                encrypted=True
            )

            response_data['mappingAvailable'] = True
            response_data['mappingId'] = session_id
            response_data['totalMappings'] = len(global_mapping.mappings)

        except Exception as e:
            print(f"ERROR saving encrypted mapping: {str(e)}")
            audit_logger.log_anonymize_error("mapping.json", str(e))

    return jsonify(response_data)


@processing_bp.route('/excel-preview/<file_id>')
def excel_preview(file_id):
    """
    Geef Excel preview: alle sheets met kolomheaders en eerste 5 rijen per sheet

    Returns:
        {
            'sheets': [
                {
                    'name': 'Sheet1',
                    'index': 0,
                    'rows': 100,
                    'columns': ['A', 'B', 'C'],
                    'preview': [['row1_a', 'row1_b'], ...]  # First 5 rows
                },
                ...
            ],
            'default_sheet': 0
        }
    """
    session_id = session.get('session_id')
    if not session_id:
        return jsonify({'error': 'Geen actieve sessie'}), 400

    session_dir = current_app.config['UPLOAD_FOLDER'] / session_id

    # Vind het bestand
    input_file = None
    for filepath in session_dir.glob(f"{file_id}.*"):
        if filepath.suffix.lower() in ['.xlsx', '.csv']:
            input_file = filepath
            break

    if not input_file:
        return jsonify({'error': 'Excel/CSV bestand niet gevonden'}), 404

    try:
        # Handle CSV files separately
        if input_file.suffix.lower() == '.csv':
            import csv

            with open(input_file, 'r', encoding='utf-8-sig', newline='') as csvfile:
                reader = csv.reader(csvfile)

                # Read all rows (header + data)
                rows = list(reader)

                if len(rows) == 0:
                    return jsonify({'error': 'CSV bestand is leeg'}), 400

                # First row is header
                columns = rows[0]

                # Preview is next 5 rows (or less if file is shorter)
                preview_rows = rows[1:6]

                sheets_info = [{
                    'name': 'CSV Data',
                    'index': 0,
                    'rows': len(rows) - 1,  # Exclude header
                    'columns': columns,
                    'preview': preview_rows
                }]

                return jsonify({
                    'success': True,
                    'filename': input_file.name,
                    'sheets': sheets_info,
                    'default_sheet': 0
                })

        # Handle Excel files with openpyxl
        from openpyxl import load_workbook

        # Load workbook in read-only mode for memory efficiency
        wb = load_workbook(input_file, read_only=True)

        sheets_info = []
        for idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]

            # Get dimensions
            max_row = ws.max_row
            max_col = ws.max_column

            # Extract column names (first row)
            columns = []
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first_row:
                columns = [str(cell) if cell is not None else f"Column {i+1}"
                          for i, cell in enumerate(first_row)]

            # Extract preview (first 5 data rows - rows 2-6)
            preview = []
            for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
                preview.append([str(cell) if cell is not None else '' for cell in row])

            sheets_info.append({
                'name': sheet_name,
                'index': idx,
                'rows': max_row - 1 if max_row > 0 else 0,  # Exclude header row
                'columns': columns,
                'preview': preview
            })

        return jsonify({
            'success': True,
            'filename': input_file.name,
            'sheets': sheets_info,
            'default_sheet': 0
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Fout bij het lezen van Excel/CSV: {str(e)}'}), 500
